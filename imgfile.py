import requests
from bs4 import BeautifulSoup
import urllib.request
# from selenium import webdriver
import time
# from MyId import Id, Pw
import pandas as pd
from pandas import DataFrame
from openpyxl import Workbook,cell
import openpyxl

def img_get_down():
   
    html = requests.get("https://www.google.com/search?q=%EB%A7%9D%EC%B4%88&tbm=isch&ved=2ahUKEwjX9dLqy73rAhWYDd4KHYpKAOYQ2-cCegQIABAA&oq=%EB%A7%9D%EC%B4%88&gs_lcp=CgNpbWcQA1AAWABg--IBaABwAHgAgAEAiAEAkgEAmAEAqgELZ3dzLXdpei1pbWc&sclient=img&ei=T89IX5eSNpib-AaKlYGwDg&bih=1022&biw=2114&hl=ko").text
    soup = BeautifulSoup(html, 'html.parser')
    div = soup.findAll(class_="t0fcAb")
    n=22
    for i in div:
        print(i)
        img_url = i.attrs["src"]

        print(img_url)
        urllib.request.urlretrieve(img_url, str(n)+".jpg")
        n=n+1

def data_get_news_do():
    
    data_url =[]
    data_title = []
    data_dict = {}
    data_date = []

    search_key = ["도시공사 디지털 뉴딜 공공데이터"]
    wb = Workbook()
    for texet in search_key:
        for i in range(0, 1):
            num = str(i*10+1)
            html = requests.get("https://search.naver.com/search.naver?&where=news&query="+texet+"&sm=tab_pge&sort=0&photo=0&field=0&reporter_article=&pd=0&ds=&de=&docid=&nso=so:r,p:all,a:all&mynews=0&cluster_rank=0&start="+num+"&refresh_start=0").text   
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.find(class_="type01")
            datas = title.find_all(class_="_sp_each_title")
            # dates = title.find_all(id="#text")
           
            for i in title.find_all(class_="txt_inline"):
                print((i.text).split(" "))
            for data in datas:
                data_title.append(data.attrs['title'])
                data_url.append(data.attrs['href'])
        #     for date in dates:
        #         data_date.append(data.text)

        # print(data_date)
        ws = wb.create_sheet(texet)
        print(ws)
        for row in range(1,len(data_title)):
            ws.cell(row=row,column=1).value = data_title[row]
            ws.cell(row=row,column=2).value = data_url[row]
            # ws.cell(row=row,column=3).value = data_date[row]
        
        data_title=[]
        data_url=[]
        data_date=[]

    wb.save("data2.xlsx")

data_get_news_do()